import sys
sys.path.append("..")

from .task import BaseTask

from mercs import Mercs
import openpyxl
import os
from uuid import uuid4
from openpyxl import load_workbook

import numpy as np
import pandas as pd
from sklearn.preprocessing import LabelEncoder

from state_manager import State, Prediction, Coordinate

import csv
import openpyxl


class MERCSTask(BaseTask):
    def __init__(self, state, context, model=None, train=[], query=None):
        super().__init__(state, context)
        self.mercs_path = ""
        self.model = model
        self.train = train
        self.query = query

    def get_train_range(self):
        selected_range = openpyxl.worksheet.cell_range.CellRange(self.context["selection"])
        table_range = None
        for t in self.state.tables:
            table_range_temp = openpyxl.worksheet.cell_range.CellRange(t.range.range_address)
            # If the selected range is in a table, we use that table as the relevant one
            if selected_range.issubset(table_range_temp):
                table_range = table_range_temp
                break
        return table_range

    def do(self):
        # Do the required actions (train, query, storing model...)
       
        # inputs from synth
        xlsx_fn = convert_csv_to_xlsx(self.state.filepath)

        xl_range_obj = self.get_train_range()
        xl_range = None
        if xl_range_obj:
            xl_range = xl_range_obj.coord
        if not xl_range:
            print("Range outside of table")
            return

        wb = load_workbook(xlsx_fn)

        empty_rows = get_empty_rows(wb, xl_range_obj)

        df, encoders, nominal_ids = extract_data(wb, xl_range, empty_rows)

        mod = Mercs(evaluation_algorithm="dummy", selection_algorithm="random", nb_iterations=3, fraction_missing=[0, 0.2])
        mod.fit(df.values, nominal_ids=nominal_ids )

        q_code = np.zeros(df.shape[1], dtype=int)

        empty_cols = get_empty_columns(wb, xl_range_obj)
        for c in empty_cols:
            q_code[c] = 1

        df_pred, _, _ = extract_data(wb, xl_range, [])

        y_pred = mod.predict(df_pred.values[:, :-1], q_code=q_code, prediction_algorithm='mi')

        predictions = []
        provenance = ("MERCS", str(uuid4()))
        for row in empty_rows:
            for col_index, col in enumerate(empty_cols):
                value = y_pred[row][col_index].item()
                if encoders[col]:
                    value = encoders[col].inverse_transform([value])[0]
                predictions.append(
                        Prediction(
                            Coordinate(col+xl_range_obj.min_col-1, row+xl_range_obj.min_row-1),
                            value,
                            1,
                            provenance,
                        )
                    )
        return self.state.add_objects(predictions)

        # print(encoders[prediction_index].inverse_transform(y_pred))

    def undo(self):
        # Undo the action (might not always be relevant)
        pass

    def description(self):
        return "MERCS prediction"


def convert_csv_to_xlsx(csv_filename, xlsx_filename=None):
    wb = openpyxl.Workbook()
    ws = wb.active

    with open(csv_filename) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)
            
    if xlsx_filename is None:
        base, _ = os.path.splitext(csv_filename)
        xlsx_filename = "{}.{}".format(base, 'xlsx')

    wb.save(xlsx_filename)
    return xlsx_filename

def get_empty_columns(wb, xl_range):
    empty_columns = set()
    ws = wb.active
    ws_range = ws[xl_range.coord]
    for row in ws_range:
        for cell in row:
            if cell.value is None:
                empty_columns.add(cell.column-xl_range.min_col)

    return list(empty_columns)

def get_empty_rows(wb, xl_range):
    empty_rows = set()
    ws = wb.active
    ws_range = ws[xl_range.coord]
    for row in ws_range:
        for cell in row:
            if cell.value is None:
                empty_rows.add(cell.row-xl_range.min_row)

    return list(empty_rows)


def extract_data(wb, xl_range, empty_rows):
    ws = wb.active
    ws_range = ws[xl_range]
    data = _parse_worksheet_range(ws_range, empty_rows)
    all_data = _parse_worksheet_range(ws_range, [])

    df_all_data = pd.DataFrame(all_data)
    df = pd.DataFrame(data)
    
    # Type-hack
    df = _convert_columns(df)
    df_all_data = _convert_columns(df_all_data)
    
    nans = df.isnull()
    
    nominal_ids = _get_nominal_ids(df)

    # Encode   
    encoders = _encoders_for_nominal(df_all_data)
    encoders_normal = _encoders_for_nominal(df)
    
    for enc, col in zip(encoders, df.columns):
        if enc:
            df[col] = enc.transform(df[col].values)  
            
    df[nans] = np.nan
    
    return df, encoders, nominal_ids

def _parse_worksheet_range(ws_range, empty_rows):
    data = []
    for index, row in enumerate(ws_range):
        if not index in empty_rows:
            data.append([cell.value for cell in row])
        
    return data

def _convert_columns(df):
    types_to_try = [int, float, "category"]
    for col in df.columns:
        for t in types_to_try:
            try:
                df[col] = df[col].astype(t, copy=False)
                break
            except:
                pass
            
    return df

def _encoders_for_nominal(df):
    encoders = []
    for col in df.columns:
        if pd.api.types.is_categorical_dtype(df[col]):
            try:
                e = LabelEncoder().fit(df[col].values)
            except:
                # Probably when a NaN is in this column.
                fix_nan_in_categorical(df[col])
                e = LabelEncoder().fit(df[col].values)
        else:
            e = None
        encoders.append(e)
    return encoders

def fix_nan_in_categorical(s):
    nans = s.isnull().values
    assert np.sum(nans) > 0
    s.cat.add_categories(['NaN'], inplace=True)
    s[nans] = 'NaN'
    return

def _get_nominal_ids(df):
    nominal_ids = [pd.api.types.is_categorical_dtype(df[col]) for col in df.columns]
    nominal_ids = [idx for idx,v in enumerate(nominal_ids) if v]
    return nominal_ids
