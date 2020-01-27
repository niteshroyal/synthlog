if not __name__ == "__main__":
    from .task import BaseTask
    from state_manager import (
        ObjectFormatting,
        FillFormatting,
        MetadataPropObject,
        Cell,
        Selection,
    )
else:

    class BaseTask:
        pass


from pyswip import Prolog, Variable
from uuid import uuid4

# StateConverter imports
import pandas as pd
import csv
import os
import openpyxl


# TODO: find a better place for this file (share it with prediction and MERCS?)
class StateParser:
    def __init__(self, state, context):
        self.state = state
        self.context = context
        self.workbook = None
        self.table_ranges = {}

    def get_cells(self, selection, color, template):
        ws = self.workbook.active
        selected_cells = []
        for table in self.table_ranges:
            range_obj = self.table_ranges[table]

            x = 0
            for row in ws[range_obj.coord]:
                if x in selection[table]:
                    selected_cells += self.get_template_cells(
                        row, color, template[table]
                    )
                x += 1
        return selected_cells

    @staticmethod
    def get_template_cells(row, color, template):
        x = 0
        i = 0
        res = []
        for cell in row:
            while template[i] < x and i + 1 < len(template):
                i += 1
            if template[i] < x and i + 1 == len(template):
                break
            elif template[i] == x:
                res.append(
                    Cell(
                        cell_address=str(cell.coordinate),
                        value=None,
                        formatting=ObjectFormatting(
                            fill=FillFormatting(color=color), font=None, borders=None
                        ),
                        metadata=[],
                    )
                )
            x += 1
        return res

    def to_dataframes(self):
        xlsx = self.__create_xlsx()
        self.workbook = openpyxl.load_workbook(xlsx)
        tables = {}
        x = 0
        for table in self.state.tables:
            tables["t" + str(x)] = StateParser.__extract_data(
                self.workbook, table.range.range_address
            )
            self.table_ranges["t" + str(x)] = openpyxl.worksheet.cell_range.CellRange(
                table.range.range_address
            )
            x += 1
        return tables

    def to_colors(self):
        """
        Has to be run after to_dataframes (else no tables are "existing")
        :return:
        """
        colors = {}
        for range_string in self.context["formats"]:
            range_excel = openpyxl.worksheet.cell_range.CellRange(
                range_string.split("!")[1]
            )  # TODO: Cleaner split
            color = self.context["formats"][range_string]["fill"]["color"]
            if color != "#FFFFFF":
                for table in self.table_ranges:
                    try:
                        intersection = self.table_ranges[table].intersection(
                            range_excel
                        )
                        if color not in colors:
                            colors[color] = {}
                        if table not in colors[color]:
                            colors[color][table] = (set(), set())
                        for i in range(
                            intersection.min_row - self.table_ranges[table].min_row,
                            intersection.max_row - self.table_ranges[table].min_row + 1,
                        ):
                            colors[color][table][0].add(i)
                        for i in range(
                            intersection.min_col - self.table_ranges[table].min_col,
                            intersection.max_col - self.table_ranges[table].min_col + 1,
                        ):
                            colors[color][table][1].add(i)
                    except ValueError:
                        pass
        return colors

    def __create_xlsx(self, xlsx_filename=None):
        csv_filename = self.state.filepath
        wb = openpyxl.Workbook()
        ws = wb.active

        with open(csv_filename) as f:
            reader = csv.reader(f, delimiter=",")
            for row in reader:
                ws.append(row)

        if xlsx_filename is None:
            base, _ = os.path.splitext(csv_filename)
            xlsx_filename = "{}.{}".format(base, "xlsx")

        wb.save(xlsx_filename)
        return xlsx_filename

    @staticmethod
    def __extract_data(wb, xl_range):
        ws = wb.active

        range_obj = openpyxl.worksheet.cell_range.CellRange(xl_range)
        range_obj.shift(row_shift=-1)
        header = []
        for row in ws[range_obj.coord]:
            header = [cell.value for cell in row]
            break

        ws_range = ws[xl_range]
        data = StateParser.__parse_worksheet_range(ws_range)
        # TODO: add headers
        df = pd.DataFrame(data, columns=header)
        return StateParser.__convert_columns(df)

    @staticmethod
    def __convert_columns(df):
        types_to_try = [int, float, "category"]
        for col in df.columns:
            for t in types_to_try:
                try:
                    df[col] = df[col].astype(t, copy=False)
                    break
                except:
                    pass
        return df

    @staticmethod
    def __parse_worksheet_range(ws_range):
        data = []
        for row in ws_range:
            data.append([cell.value for cell in row])
        return data


class ValueSet:
    """
    Used to encode strings as integers for subtle
    """

    def __init__(self):
        self.values = {}
        self.reverse = []

    def get_value(self, key):
        if key not in self.values:
            return None
        return self.values[key]

    def store_value(self, key):
        if key in self.values:
            return self.values[key]
        self.values[key] = len(self.reverse)
        self.reverse.append(key)
        return self.values[key]


class BaseSelectionTask(BaseTask):
    def do(self):
        self.init()
        table_colors, templates = self.run()
        cells = self.converter.get_cells(table_colors, self.relevant_color, templates)
        cells += self.converter.get_cells(
            {table: self.irrelevant[0] for table in self.irrelevant},
            self.irrelevant_color,
            templates,
        )
        selections = [
            Selection(cell=c, provenance=(self.provenance, str(uuid4()))) for c in cells
        ]

        return self.state.add_objects(selections)

    def description(self) -> str:
        return "LGG Selection"

    def is_available(self) -> bool:
        colors = set()
        for range_string in self.context["formats"]:
            color = self.context["formats"][range_string]["fill"]["color"]
            if color != "#FFFFFF":
                colors.add(color)
                if len(colors) > 1:
                    return True
        return False

    def __init__(self, state, context: dict):
        super().__init__(state, context)
        self._irrelevant_tryout = 0
        self.values = None
        self.tables = None
        self.relevant = None
        self.irrelevant = None
        self.subtle_path = "./src/server/resources/tasks/prolog/subtle-2.2.pl"
        self.glgg_path = "./src/server/resources/tasks/prolog/glgg.pl"
        self.provenance = "LGG_selection"

    #######################
    #                     #
    #       Methods       #
    #                     #
    #######################

    def init(self):
        self._irrelevant_tryout = 0
        self.values = ValueSet()
        self.converter = StateParser(self.state, self.context)
        (
            self.tables,
            self.relevant,
            self.irrelevant,
            self.relevant_color,
            self.irrelevant_color,
        ) = self.extract_parameters_from_state()

    def run(self):
        prolog_examples, templates, irrelevant_tuples = self.build_data()
        model = self.build_prolog_model(prolog_examples, templates, irrelevant_tuples)

        table_colors = {}
        for res in model.query("relevant(X, Y)"):
            if res["X"] not in table_colors:
                table_colors[res["X"]] = set()
            table_colors[res["X"]].add(res["Y"])
        return table_colors, templates

    def assertz_prolog_tuples(self, prolog, templates):
        for key in self.tables:
            for i, row in self.tables[key].iloc[:, templates[key]].iterrows():
                irow = [i] + list(row)
                prolog.assertz(self.build_prolog_tuple(key, irow))

    def build_data(self):
        tuples, templates = self.build_tuples(self.relevant)
        irrelevant_tuples, _ = self.build_tuples(self.irrelevant)
        examples = self.build_examples(tuples, irrelevant_tuples)
        prolog_examples = self.build_prolog_examples(examples)
        return prolog_examples, templates, irrelevant_tuples

    def build_prolog_model(self, prolog_examples, templates, irrelevant_tuples):
        model = self.build_model(prolog_examples)
        self.assertz_prolog_tuples(model, templates)
        relevant_rules = self.lggs_to_rules(
            self.golem(model, prolog_examples, irrelevant_tuples)
        )
        self.assertz_rules(model, relevant_rules)
        return model

    def build_examples(self, tuples, irrelevant):
        labels = list(tuples.keys())
        joins = self.detect_join_order(tuples)
        return self.rec_build_examples([], labels, tuples, joins, irrelevant)

    def rec_build_examples(self, examples, labels, tuples, joins, irrelevant):
        if not joins:
            return examples
        nexamples = examples
        left = labels[joins[0][0]]
        if not nexamples:
            for i, t in tuples[left].iterrows():
                nexamples.append({left: t})
        if nexamples:
            if left in nexamples[0]:
                right = labels[joins[0][1]]
                if right not in nexamples[0]:
                    shared_columns = joins[0][2]
                    extended_examples = self.extend_examples(
                        nexamples, tuples, left, right, shared_columns, irrelevant
                    )
                    njoins = joins[1:] if len(joins) > 1 else []
                    return self.rec_build_examples(
                        extended_examples, labels, tuples, njoins, irrelevant
                    )
        return []

    def extend_examples(
        self, examples, tuples, left, right, shared_columns, irrelevant
    ):
        extended_examples = []
        for example in examples:
            right_df = tuples[right]
            for c in shared_columns:
                right_df = right_df[right_df[c] == example[left][c]]
            if right_df.shape[0] == 0:
                right_df = self.extend_example(
                    right,
                    example[left],
                    tuples[right].columns,
                    shared_columns,
                    irrelevant,
                )
            nexample = example
            nexample[right] = right_df
            extended_examples.append(nexample)
        return extended_examples

    def detect_join_order(self, tuples):
        joins = self.detect_joins(tuples)
        return self.rec_detect_join_order(joins)

    def rec_detect_join_order(self, joins):
        if not joins:
            return []
        for i in range(len(joins)):
            first = True
            for j in range(len(joins)):
                if joins[i][0] == joins[j][1] or (
                    joins[i][1] == joins[j][1] and len(joins[i][2]) < len(joins[j][2])
                ):
                    first = False
                    break
            if first:
                if i == len(joins) - 1:
                    return [joins[i]] + self.rec_detect_join_order(joins[:-1])
                return [joins[i]] + self.rec_detect_join_order(
                    joins[:i] + joins[i + 1 :]
                )
        return joins

    def build_model(self, examples):
        pl = Prolog()
        pl.consult(self.subtle_path)
        pl.consult(self.glgg_path)
        for example in examples:
            pl.assertz(example[:-1])
        return pl

    def build_prolog_examples(self, examples, store=True):
        facts = []
        x = 0
        for example in examples:
            fact = "example(" + str(x) + ", ["
            example_lst = []
            for key in example.keys():
                if len(example[key].shape) == 1:
                    example_lst.append(
                        self.build_prolog_tuple(key, example[key], store)
                    )
                else:
                    for i, row in example[key].iterrows():
                        example_lst.append(self.build_prolog_tuple(key, row, store))
            fact += ",".join(example_lst) + "])."
            facts.append(fact)
            x += 1
        return facts

    def build_prolog_tuple(self, label, row, store=True):
        return label + "(" + ",".join([self.prolog_str(x, store) for x in row]) + ")"

    def build_tuples(self, color):
        tuples = {}
        templates = {}
        for label in self.tables:
            rows = []
            columns = []
            if label in color:
                rows = sorted(list(color[label][0]))
                columns = sorted(list(color[label][1]))
            tuples[label] = self.tables[label].iloc[rows, columns]
            templates[label] = columns
        return tuples, templates

    def extend_example(self, label, example, columns, shared_columns, irrelevant):
        other = self.tables[label]
        for c in shared_columns:
            other = other[other[c] == example[c]]
        return other[~other.index.isin(irrelevant[label].index)][columns]

    def golem(self, prolog, examples, irrelevants):
        res, _ = self.rec_golem(prolog, examples, irrelevants, [], 0)
        return res

    def extract_parameters_from_state(self):
        dfs = self.converter.to_dataframes()
        colors = self.converter.to_colors()
        if len(colors.keys()) < 2:
            raise ValueError("Too few colors are given in input (need at least 2)")
        relevant_color = list(colors.keys())[0]
        irrelevant_color = list(colors.keys())[1]
        relevant = colors[relevant_color]
        irrelevant = colors[irrelevant_color]
        return dfs, relevant, irrelevant, relevant_color, irrelevant_color

    def prolog_str(self, string, store=False):
        if store:
            if isinstance(string, int):
                return str(string)
            return str(self.values.store_value(string))
        return str(string)

    def is_irrelevant(self, prolog, lgg, irrelevants):
        label = str(self._irrelevant_tryout)
        self._irrelevant_tryout += 1
        relevant_rules = self.lggs_to_rules([lgg], False, label)

        self.assertz_rules(prolog, relevant_rules)
        for res in prolog.query("relevant" + label + "(X, Y)"):
            if res["X"] in irrelevants and res["Y"] in irrelevants[res["X"]].index:
                return True

        return lgg is None

    def lggs_to_rules(self, lggs, store=False, predicat_label=""):
        relevant_rules = []
        for lgg in lggs:
            table_ids = {}
            body = self.lgg_to_rule_body(lgg, table_ids, store)
            for label in table_ids:
                for tid in table_ids[label]:
                    relevant_rules.append(
                        "relevant"
                        + predicat_label
                        + "("
                        + str(label)
                        + ", I"
                        + str(tid)
                        + ") :- "
                        + ",".join(body)
                        + "."
                    )
        return relevant_rules

    def lgg_to_rule_body(self, lgg, table_ids, store=False):
        body = []
        body_variables = {}
        ids = 0

        for fact in lgg:
            fact_lst = ["I" + str(ids)]
            if fact.name not in table_ids:
                table_ids[fact.name] = []
            table_ids[fact.name].append(ids)
            ids += 1
            for arg in fact.args:
                if isinstance(arg, Variable):
                    sarg = str(arg)
                    if sarg not in body_variables:
                        body_variables[sarg] = len(body_variables)
                    fact_lst.append("X" + str(body_variables[sarg]))
                else:
                    fact_lst.append(self.prolog_str(arg, store))
            body.append(str(fact.name) + "(" + ",".join(fact_lst) + ")")
        return body

    def rec_golem(self, prolog, examples, irrelevants, indices, pivot):
        result = []
        indices_set = []
        for i in range(pivot, len(examples)):
            nindices = indices + [i]
            if not BaseSelectionTask.in_indices_set(indices_set, nindices):
                lgg = None
                for res in prolog.query("list_lgg(" + str(nindices) + ", X)"):
                    lgg = res["X"]
                if not self.is_irrelevant(prolog, lgg, irrelevants):
                    children_res, children_indices = self.rec_golem(
                        prolog, examples, irrelevants, nindices, i + 1
                    )
                    if children_res:
                        indices_set += children_indices
                        result += children_res
                    else:
                        indices_set.append(set(nindices))
                        result.append(lgg)
        return result, indices_set

    #######################
    #                     #
    #       Statics       #
    #                     #
    #######################

    @staticmethod
    def assertz_rules(prolog, rules):
        for rule in rules:
            prolog.assertz(rule[:-1])

    @staticmethod
    def detect_joins(tuples):
        pair_join = []
        labels = list(tuples.keys())
        for i in range(len(labels)):
            for j in range(i + 1, len(labels)):
                shared_columns = [
                    x for x in tuples[labels[i]] if x in tuples[labels[j]]
                ]
                if shared_columns:
                    if tuples[labels[i]].loc[:, shared_columns].duplicated().any():
                        pair_join.append((j, i, shared_columns))
                    else:
                        pair_join.append((i, j, shared_columns))
        return pair_join

    @staticmethod
    def in_indices_set(indices_set, indices):
        for s in indices_set:
            if set(indices) <= s:
                return True
        return False


class MultipleSelectionTask(BaseSelectionTask):
    def __init__(self, state, context: dict):
        super().__init__(state, context)

    def description(self) -> str:
        return "LGG clustering selection"


class TestSelectionTask(BaseSelectionTask):
    def __init__(self, tables, colors):
        self.tables = tables
        self.colors = colors
        self._irrelevant_tryout = 0
        self.values = ValueSet()

        self.relevant = colors["relevant"]
        self.irrelevant = colors["irrelevant"]
        self.relevant_color = "relevant"
        self.irrelevant_color = "irrelevant"

        self.subtle_path = "./src/server/resources/tasks/prolog/subtle-2.2.pl"
        self.glgg_path = "./src/server/resources/tasks/prolog/glgg.pl"

    def test(self):
        table_colors, templates = self.run()
        print(table_colors)


def test():
    tables = {
        "sales": pd.DataFrame(
            [
                ["Vanilla", "Florence", 610, 190, 670, 1470, "YES"],
                ["Banana", "Stockholm", 170, 690, 520, 1380, "YES"],
                ["Chocolate", "Copenhagen", 560, 320, 140, 1020, "YES"],
                ["Banana", "Berlin", 610, 640, 320, 1570, "NO"],
                ["Stracciatella", "Florence", 300, 270, 290, 860, "NO"],
                ["Chocolate", "Milan", 430, 350, "?", "?", "?"],
                ["Banana", "Aachen", 250, 650, "?", "?", "?"],
                ["Chocolate", "Brussels", 210, 280, "?", "?", "?"],
            ],
            columns=["Type", "City", "June", "July", "Aug", "Total", "Profit"],
        ),
        "providers": pd.DataFrame(
            [
                ["Vanilla", "Florence", 1, "Cheap", "Bad"],
                ["Vanilla", "Florence", 2, "Regular", "Good"],
                ["Stracciatella", "Florence", 1, "Regular", "Great"],
                ["Chocolate", "Copenhagen", 3, "Cheap", "Good"],
                ["Chocolate", "Milan", 4, "Regular", "Good"],
                ["Chocolate", "Milan", 5, "Expensive", "Great"],
                ["Chocolate", "Brussels", 6, "Regular", "Good"],
                ["Chocolate", "Brussels", 6, "Expensive", "Good"],
            ],
            columns=["Type", "City", "ProviderID", "Price", "Quality"],
        ),
    }

    colors = {
        "relevant": {
            "sales": ({0, 2, 5, 7}, {0, 1, 2, 3, 4, 6}),
            "providers": ({0, 1, 3}, {0, 1, 3, 4}),
        },
        "irrelevant": {"sales": ({}, {}), "providers": ({5, 7}, {4})},
    }

    TestSelectionTask(tables, colors).test()


if __name__ == "__main__":
    test()
